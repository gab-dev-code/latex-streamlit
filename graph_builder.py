import pandas as pd # type: ignore
import matplotlib.pyplot as plt # type: ignore
import numpy as np # type: ignore
from openpyxl import load_workbook # type: ignore
import tempfile
import shutil

excel_source = "data_test.xlsx"

def generate_dual_iv_curves(excel_source, series_name, output_path="photos/IV_curve.png", voc_input=None, isc_input=None, pmax_input=None):
    """
    Generate I-V curves for a specific panel series.
    
    Parameters:
    -----------
    excel_source : str or file-like
        Path to Excel file or file object
    series_name : str
        Name of the series to generate curves for
    voc_input : float, optional
        Open circuit voltage (will be written to D18 in sheet 2)
    isc_input : float, optional
        Short circuit current (will be written to D21 in sheet 2)
    pmax_input : float, optional
        Maximum power (will be written to D24 in sheet 2)
    """
    
    # We need a writable file for openpyxl to work
    # Always create a temp copy since uploads directory is read-only
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    
    # Copy the source to temp file
    if hasattr(excel_source, 'read'):
        # It's a file object
        excel_source.seek(0)
        shutil.copyfileobj(excel_source, temp_file)
        excel_source.seek(0)
    else:
        # It's a file path
        with open(excel_source, 'rb') as source:
            shutil.copyfileobj(source, temp_file)
    
    temp_file.close()
    working_file = temp_file.name
    
    try:
        # Update Sheet 2 with the values from Sheet 1
        wb = load_workbook(working_file)
        ws = wb.worksheets[1]  # Sheet 2 (0-indexed)
        
        # Update cells if values are provided
        if voc_input is not None:
            ws['D18'] = float(voc_input)
            print(f"Updated D18 (Voc) = {voc_input}")
        
        if series_name is not None:
            ws['D19'] = str(series_name)
            print(f"Updated D19 (Series) = {series_name}")
        
        if isc_input is not None:
            ws['D21'] = float(isc_input)
            print(f"Updated D21 (Isc) = {isc_input}")
        
        if pmax_input is not None:
            ws['D24'] = float(pmax_input)
            print(f"Updated D24 (Pmax) = {pmax_input}")
        
        # Save the workbook
        wb.save(working_file)
        wb.close()
        
        # Now read the updated data with pandas
        df_raw = pd.read_excel(working_file, sheet_name=1, header=None)
        
        # Read Voc and Isc directly from column D (index 3) since we just wrote them there
        voc = float(df_raw.iloc[17, 3])  # D18
        isc = float(df_raw.iloc[20, 3])  # D21
        
        print(f"Reading back: Voc={voc}, Isc={isc}")
        
        if pd.isna(voc) or pd.isna(isc):
            print(f"Error: Invalid values - Voc={voc}, Isc={isc}")
            return False
        
    except Exception as e:
        print(f"Data extraction error: {e}")
        import traceback
        traceback.print_exc()
        return False
    finally:
        # Clean up temp file
        import os
        try:
            os.unlink(temp_file.name)
        except:
            pass

    # 4. Create the Figure
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(10, 4))

    # --- Graph 1: Irradiance ---
    v = np.linspace(0, voc, 100)
    for G, col in zip([1000, 800, 600, 400, 200], ['#002060', '#0070C0', '#00B0F0', '#92D050', '#FFFF00']):
        # Simple PV Curve Model: I = Isc*(G/1000) - alpha*(exp(V/beta)-1)
        i_curr = (isc * (G/1000)) * (1 - (v/voc)**4) 
        ax1.plot(v, i_curr, label=f'{G} W/m²', color=col)
    
    ax1.set_title("I-V characteristics at different irradiances")
    ax1.set_xlabel("Voltage (V)")
    ax1.set_ylabel("Current (A)")
    ax1.legend(fontsize='small')

    # --- Graph 2: Temperature ---
    for T, col in zip([70, 50, 25, 0], ['#C00000', '#FFC000', '#0070C0', '#002060']):
        # Voc drops with temp (~0.3%/C), Isc rises slightly (~0.05%/C)
        v_t = voc * (1 - 0.003 * (T - 25))
        i_t = isc * (1 + 0.0005 * (T - 25))
        v_range = np.linspace(0, v_t, 100)
        i_curr = i_t * (1 - (v_range/v_t)**4)
        ax2.plot(v_range, i_curr, label=f'{T} °C', color=col)

    ax2.set_title("I-V characteristics at different temperatures")
    ax2.set_xlabel("Voltage (V)")
    ax2.set_ylabel("Current (A)")
    ax2.legend(fontsize='small')

    plt.tight_layout()
    plt.savefig(output_path, dpi=300)
    plt.close()
    
    print(f"Graph saved to {output_path}")
    return True